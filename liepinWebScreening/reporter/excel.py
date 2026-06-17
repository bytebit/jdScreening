"""
Excel 报告生成器
================
将分析结果导出为结构化的 Excel 报告，HR 可以直接查看和操作。

报告包含：
  1. 筛选总览 Sheet — 所有候选人的汇总排名
  2. 详细评估 Sheet — 每个候选人的逐项评估
  3. (可选) 对比 Sheet — S/A 级候选人的横向对比
"""

import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import config as cfg
from models import AnalysisResult


class ExcelReporter:
    """
    Excel 报告生成器

    Example:
        reporter = ExcelReporter()
        filepath = reporter.generate("高级Java工程师", results)
        print(f"报告已保存: {filepath}")
    """

    def __init__(self):
        self.output_dir = Path(cfg.REPORT_DIR)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def generate(
        self,
        job_title: str,
        results: list[AnalysisResult],
        output_path: Optional[str] = None,
    ) -> str:
        """
        生成 Excel 筛选报告

        Args:
            job_title: 职位名称
            results: 分析结果列表（已按评分排序）
            output_path: 输出路径，None 则自动生成

        Returns:
            生成的 Excel 文件路径
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        if not output_path:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = str(
                self.output_dir / f"筛选报告_{job_title}_{timestamp}.xlsx"
            )

        wb = Workbook()

        # ── Sheet 1: 筛选总览 ──
        ws = wb.active
        ws.title = "筛选总览"
        self._write_overview_sheet(ws, job_title, results)

        # ── Sheet 2: 详细评估 ──
        ws2 = wb.create_sheet("详细评估")
        self._write_detail_sheet(ws2, results)

        # ── 保存 ──
        wb.save(output_path)
        print(f"📊 报告已导出: {output_path}")
        return output_path

    # ────────────────────────────────────────────────
    # Sheet 1: 筛选总览
    # ────────────────────────────────────────────────
    def _write_overview_sheet(self, ws, job_title: str, results: list[AnalysisResult]):
        """写入筛选总览 Sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # 样式定义
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        title_font = Font(name="微软雅黑", bold=True, size=14)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 等级颜色
        level_fills = {
            "S": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            "A": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
            "B": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
            "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "D": PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        }

        # 标题行
        ws.merge_cells("A1:H1")
        ws["A1"] = f"简历筛选报告 — {job_title}"
        ws["A1"].font = title_font
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # 统计信息
        ws.merge_cells("A2:H2")
        total = len(results)
        s_count = sum(1 for r in results if r.final_level == "S")
        a_count = sum(1 for r in results if r.final_level == "A")
        b_count = sum(1 for r in results if r.final_level == "B")
        c_count = sum(1 for r in results if r.final_level == "C")
        d_count = sum(1 for r in results if r.final_level == "D")
        ws["A2"] = f"总计: {total}人  S:{s_count}  A:{a_count}  B:{b_count}  C:{c_count}  D:{d_count}"
        ws["A2"].font = Font(name="微软雅黑", size=10)
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

        # 表头
        headers = ["排名", "姓名", "综合评分", "等级", "推荐决策", "规则筛选", "语义匹配", "摘要"]
        col_widths = [6, 12, 10, 6, 12, 12, 10, 50]

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[chr(64 + col_idx)].width = width

        # 数据行
        for rank, result in enumerate(results, 1):
            row = 4 + rank
            ws.cell(row=row, column=1, value=rank).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=2, value=result.candidate_name)
            ws.cell(row=row, column=3, value=result.final_score).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=4, value=result.final_level).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=5, value=result.final_recommendation)
            ws.cell(row=row, column=6, value="通过" if result.stage1.passed else "淘汰")

            # 语义匹配分
            sim_score = result.stage2.similarity_score
            ws.cell(row=row, column=7, value=f"{sim_score:.2%}" if sim_score > 0 else "—")
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")

            ws.cell(row=row, column=8, value=result.final_summary)

            # 行高
            ws.row_dimensions[row].height = 22

            # 等级颜色
            level_fill = level_fills.get(result.final_level)
            if level_fill:
                for col in range(1, 9):
                    ws.cell(row=row, column=col).fill = level_fill

            # 边框
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

        # 冻结前几行
        ws.freeze_panes = "A5"

    # ────────────────────────────────────────────────
    # Sheet 2: 详细评估
    # ────────────────────────────────────────────────
    def _write_detail_sheet(self, ws, results: list[AnalysisResult]):
        """写入详细评估 Sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # 只写 S/A/B 级的详细评估
        detailed = [r for r in results if r.final_level in ("S", "A", "B")]
        if not detailed:
            ws["A1"] = "无详细评估数据（所有简历均未达到B级以上）"
            return

        headers = [
            "姓名", "等级", "评分", "推荐决策",
            "优势", "风险", "面试重点", "面试建议题",
        ]
        col_widths = [12, 6, 8, 12, 40, 40, 40, 40]

        header_font = Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + col_idx)].width = width

        for row_idx, result in enumerate(detailed, 2):
            s3 = result.stage3

            ws.cell(row=row_idx, column=1, value=result.candidate_name)
            ws.cell(row=row_idx, column=2, value=result.final_level).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=3, value=result.final_score).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=4, value=result.final_recommendation)
            ws.cell(row=row_idx, column=5, value="\n".join(s3.strengths) if s3.strengths else "")
            ws.cell(row=row_idx, column=6, value="\n".join(s3.concerns) if s3.concerns else "")
            ws.cell(row=row_idx, column=7, value="\n".join(s3.interview_focus) if s3.interview_focus else "")
            ws.cell(row=row_idx, column=8, value="\n".join(s3.details.get("suggested_questions", [])) if s3.details else "")

            # 自动换行
            for col in [5, 6, 7, 8]:
                ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            ws.row_dimensions[row_idx].height = max(60, len(s3.strengths) * 20)

        ws.freeze_panes = "A2"
